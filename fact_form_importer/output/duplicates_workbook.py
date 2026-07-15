"""Generate a focused workbook for deciding how to handle duplicate form submissions."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from fact_form_importer.ingest.column_mapping import load_column_mapping
from fact_form_importer.models.court_submission import CourtSubmission
from fact_form_importer.output.duplicate_review import (
    duplicate_timestamp,
    format_source_date,
    group_duplicate_submissions,
    most_recent_duplicate_submission,
    sort_duplicate_submissions,
)

WORKBOOK_NAME = "duplicate_forms_review.xlsx"
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
CANDIDATE_FILL = PatternFill("solid", fgColor="E2F0D9")
DECISION_FILL = PatternFill("solid", fgColor="FFF2CC")
DEFAULT_COLUMN_MAPPING_PATH = Path("config/column_mapping.json")

FACILITY_LABELS = {
    "accessible_parking": "Accessible parking",
    "accessible_parking_phone": "Accessible parking booking phone",
    "accessible_toilet_description": "Accessible toilet description",
    "accessible_entrance": "Accessible entrance",
    "accessible_entrance_support_phone": "Accessible entrance support phone",
    "hearing_enhancement_equipment": "Hearing enhancement equipment",
    "lift_available": "Lift available",
    "lift_door_width": "Lift door width (cm)",
    "lift_weight_limit": "Lift weight limit (kg)",
    "quiet_room_available": "Quiet room (accessibility question)",
    "parking_available": "Parking available",
    "food_and_drink": "Food and drink",
    "separate_waiting_areas": "Separate waiting areas",
    "child_waiting_area": "Separate waiting area for children",
    "quiet_room_available_2": "Quiet room (facilities question)",
    "baby_changing": "Baby changing facilities",
    "wifi_available": "WiFi available",
}

COUNTER_SERVICE_LABELS = {
    "specific_courts": "Specific courts",
    "assists_with": "Can assist with",
    "appointment_contact": "Appointment booking contact",
    "same_monday_to_friday": "Same Monday to Friday",
}

INTERVIEW_ROOM_LABELS = {
    "has_interview_rooms": "Interview rooms available",
    "room_count": "Interview room count",
    "booking_phone": "Interview room booking phone",
}

def write_duplicate_review_workbook(
    submissions: list[CourtSubmission], output_path: Path, summary: dict[str, Any]
) -> Path:
    """Write duplicate groups with the authoritative latest-form selection."""

    output_path.mkdir(parents=True, exist_ok=True)
    groups = group_duplicate_submissions(submissions)
    workbook = Workbook()
    workbook.remove(workbook.active)

    group_rows = _group_rows(groups)
    _add_summary_sheet(workbook, summary, group_rows)
    _add_group_overview_sheet(workbook, group_rows)
    _add_duplicate_forms_sheet(workbook, group_rows)
    _add_duplicate_form_data_sheet(workbook, group_rows)
    _add_decision_log_sheet(workbook, group_rows)

    path = output_path / WORKBOOK_NAME
    workbook.save(path)
    return path


def _group_rows(groups: dict[str, list[CourtSubmission]]) -> list[dict[str, Any]]:
    rows = []
    for group_number, (court_slug, submissions) in enumerate(sorted(groups.items()), start=1):
        candidate = most_recent_duplicate_submission(submissions)
        candidate_timestamp = duplicate_timestamp(candidate) if candidate else None
        rows.append(
            {
                "group": f"D{group_number:03d}",
                "court_slug": court_slug,
                "submissions": sort_duplicate_submissions(submissions),
                "candidate": candidate,
                "candidate_timestamp": candidate_timestamp,
            }
        )
    return rows


def _add_summary_sheet(
    workbook: Workbook, summary: dict[str, Any], group_rows: list[dict[str, Any]]
) -> None:
    rows = [
        ["Duplicate form review", None],
        ["Run ID", summary.get("run_id")],
        ["Source file", summary.get("source_file")],
        ["Duplicate court groups", len(group_rows)],
        ["Duplicate form rows", sum(len(group["submissions"]) for group in group_rows)],
        [],
        [
            "How to use this workbook",
            "Audit each duplicate group in Duplicate form data. It is self-contained and does not require the NSU cleaned review workbook.",
        ],
        [
            "Authoritative latest form",
            "The importer selects Completion time, then Last modified time, then Start time, with the highest source row as the final fallback.",
        ],
        [
            "Operational effect",
            "Only the authoritative row is validated, sent to OS/LLM review, or given API actions. Older rows remain audit evidence.",
        ],
        [
            "Raw and cleaned evidence",
            "Duplicate form data shows original non-empty submitted answers alongside readable cleaned data. Source row numbers remain available if cross-checking against the NSU workbook is useful.",
        ],
    ]
    worksheet = workbook.create_sheet("Summary")
    for row in rows:
        worksheet.append(row)
    worksheet["A1"].font = HEADER_FONT
    worksheet["A1"].fill = HEADER_FILL
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 110
    worksheet.freeze_panes = "A2"


def _add_group_overview_sheet(workbook: Workbook, group_rows: list[dict[str, Any]]) -> None:
    rows = [
        [
            "duplicate_group",
            "court_slug",
            "form_count",
            "source_row_numbers",
            "authoritative_source_row",
            "candidate_timestamp",
            "candidate_timestamp_field",
            "decision_required",
            "nsu_decision",
            "selected_source_row",
            "decision_notes",
        ]
    ]
    for group in group_rows:
        candidate_timestamp = group["candidate_timestamp"]
        rows.append(
            [
                group["group"],
                group["court_slug"],
                len(group["submissions"]),
                ", ".join(str(item.source.source_row_number) for item in group["submissions"]),
                group["candidate"].source.source_row_number if group["candidate"] else None,
                candidate_timestamp.display_value if candidate_timestamp else None,
                candidate_timestamp.source_label if candidate_timestamp else "No timestamp available",
                "No - latest-form policy applied",
                "Use latest completed form",
                group["candidate"].source.source_row_number if group["candidate"] else None,
                None,
            ]
        )
    worksheet = _write_table(workbook, "Duplicate group overview", rows)
    _highlight_decision_columns(worksheet, start_row=2)


def _add_duplicate_forms_sheet(workbook: Workbook, group_rows: list[dict[str, Any]]) -> None:
    rows = [
        [
            "duplicate_group",
            "court_slug",
            "authoritative_latest_form",
            "source_row_number",
            "forms_id",
            "completion_time",
            "last_modified_time",
            "start_time",
            "comparison_timestamp",
            "comparison_timestamp_field",
            "submitter_name",
            "submitter_email",
            "raw_court_slug",
            "status",
            "issue_codes",
            "address_count",
            "contact_count",
            "opening_hours_count",
            "review_note",
        ]
    ]
    candidate_rows = set()
    for group in group_rows:
        candidate = group["candidate"]
        for submission in group["submissions"]:
            timestamp = duplicate_timestamp(submission)
            row_number = submission.source.source_row_number
            if candidate and row_number == candidate.source.source_row_number:
                candidate_rows.add(len(rows) + 1)
            rows.append(
                [
                    group["group"],
                    group["court_slug"],
                    "Yes - authoritative" if submission is candidate else "No - superseded",
                    row_number,
                    submission.source.forms_id,
                    _display_source_value(submission.source.completion_time),
                    _display_source_value(submission.source.last_modified_time),
                    _display_source_value(submission.source.start_time),
                    timestamp.display_value if timestamp else None,
                    timestamp.source_label if timestamp else "No timestamp available",
                    submission.source.submitter_name,
                    submission.source.submitter_email,
                    submission.court_slug_raw,
                    submission.status,
                    ", ".join(sorted({issue.code for issue in submission.issues})),
                    len(submission.addresses),
                    len(submission.contacts),
                    len(submission.opening_hours),
                    "Authoritative latest form" if submission is candidate else "Superseded audit row",
                ]
            )
    worksheet = _write_table(workbook, "Duplicate form rows", rows)
    for row_number in candidate_rows:
        for cell in worksheet[row_number]:
            cell.fill = CANDIDATE_FILL


def _add_duplicate_form_data_sheet(workbook: Workbook, group_rows: list[dict[str, Any]]) -> None:
    """Add all useful form content so duplicate decisions do not need a second workbook."""

    rows = [
        [
            "duplicate_group",
            "court_slug",
            "authoritative_latest_form",
            "source_row_number",
            "forms_id",
            "completion_time",
            "last_modified_time",
            "start_time",
            "raw_court_slug",
            "cleaned_court_slug",
            "submitted_form_answers",
            "cleaned_facilities",
            "cleaned_translation_services",
            "cleaned_addresses",
            "cleaned_counter_service",
            "cleaned_interview_rooms",
            "cleaned_contact_details",
            "cleaned_court_opening_hours",
            "status",
            "issue_codes",
        ]
    ]
    candidate_rows = set()
    raw_labels = _raw_field_labels(group_rows)

    for group in group_rows:
        candidate = group["candidate"]
        for submission in group["submissions"]:
            row_number = submission.source.source_row_number
            if candidate and row_number == candidate.source.source_row_number:
                candidate_rows.add(len(rows) + 1)
            rows.append(
                [
                    group["group"],
                    group["court_slug"],
                    "Yes - authoritative" if submission is candidate else "No - superseded",
                    row_number,
                    submission.source.forms_id,
                    _display_source_value(submission.source.completion_time),
                    _display_source_value(submission.source.last_modified_time),
                    _display_source_value(submission.source.start_time),
                    submission.court_slug_raw,
                    submission.court_slug,
                    _format_raw_answers(submission, raw_labels),
                    _format_key_values(submission.facilities, FACILITY_LABELS),
                    _format_translation_services(submission),
                    _format_addresses(submission),
                    _format_counter_service(submission),
                    _format_key_values(submission.interview_rooms, INTERVIEW_ROOM_LABELS),
                    _format_contacts(submission),
                    _format_opening_hours(submission),
                    submission.status,
                    ", ".join(sorted({issue.code for issue in submission.issues})),
                ]
            )

    worksheet = _write_table(
        workbook,
        "Duplicate form data",
        rows,
        wrap_text=True,
        data_row_height=120,
    )
    _set_column_widths(
        worksheet,
        {
            "A": 16,
            "B": 30,
            "C": 26,
            "D": 19,
            "E": 22,
            "F": 20,
            "G": 22,
            "H": 20,
            "I": 28,
            "J": 28,
            "K": 60,
            "L": 46,
            "M": 34,
            "N": 60,
            "O": 50,
            "P": 38,
            "Q": 60,
            "R": 60,
            "S": 28,
            "T": 40,
        },
    )
    for row_number in candidate_rows:
        for cell in worksheet[row_number]:
            cell.fill = CANDIDATE_FILL


def _add_decision_log_sheet(workbook: Workbook, group_rows: list[dict[str, Any]]) -> None:
    rows = [
        [
            "duplicate_group",
            "court_slug",
            "date_based_candidate_source_row",
            "all_source_row_numbers",
            "nsu_decision",
            "selected_source_row",
            "decision_maker",
            "decision_date",
            "decision_notes",
        ]
    ]
    for group in group_rows:
        rows.append(
            [
                group["group"],
                group["court_slug"],
                group["candidate"].source.source_row_number if group["candidate"] else None,
                ", ".join(str(item.source.source_row_number) for item in group["submissions"]),
                None,
                None,
                None,
                None,
                None,
            ]
        )
    worksheet = _write_table(workbook, "Decision log", rows)
    _highlight_decision_columns(worksheet, start_row=2)


def _display_source_value(value: str | None) -> str | None:
    return format_source_date(value)


def _raw_field_labels(group_rows: list[dict[str, Any]]) -> dict[str, str]:
    """Return human-readable labels for the source columns used in populated forms."""

    if not any(submission.raw for group in group_rows for submission in group["submissions"]):
        return {}

    try:
        mapping = load_column_mapping(DEFAULT_COLUMN_MAPPING_PATH)
    except (OSError, ValueError):
        return {}

    labels: dict[str, str] = {}
    _add_column_labels(labels, mapping.scalars)
    _add_group_column_labels(labels, "Address", mapping.address_groups)
    _add_column_labels(labels, mapping.counter_service, prefix="Counter service")
    _add_column_labels(labels, mapping.interview_rooms, prefix="Interview rooms")
    _add_group_column_labels(labels, "Contact detail", mapping.contact_detail_groups)
    _add_group_column_labels(labels, "Court opening hours", mapping.opening_hours_groups)
    return labels


def _add_column_labels(
    labels: dict[str, str],
    columns: list[Any],
    prefix: str | None = None,
) -> None:
    for column in columns:
        description = column.expected_header or column.field.replace("_", " ")
        labels[column.column] = f"{prefix} - {description}" if prefix else description


def _add_group_column_labels(
    labels: dict[str, str],
    group_name: str,
    groups: list[Any],
) -> None:
    for group in groups:
        _add_column_labels(labels, group.columns, prefix=f"{group_name} {group.index}")


def _format_raw_answers(submission: CourtSubmission, labels: dict[str, str]) -> str | None:
    values = []
    for column, value in submission.raw.items():
        if column in {"A", "B", "C", "D", "E", "F"} or not _has_value(value):
            continue
        label = labels.get(column, f"Column {column}")
        values.append(f"{label}: {_display_value(value)}")
    return "\n".join(values) if values else None


def _format_key_values(values: dict[str, Any], labels: dict[str, str]) -> str | None:
    lines = [
        f"{labels.get(key, key.replace('_', ' ').capitalize())}: {_display_value(value)}"
        for key, value in values.items()
        if _has_value(value)
    ]
    return "\n".join(lines) if lines else None


def _format_translation_services(submission: CourtSubmission) -> str | None:
    values = {
        "Phone": submission.translation_phone,
        "Email": submission.translation_email,
    }
    lines = [f"{label}: {_display_value(value)}" for label, value in values.items() if _has_value(value)]
    return "\n".join(lines) if lines else None


def _format_addresses(submission: CourtSubmission) -> str | None:
    addresses = []
    for address in submission.addresses:
        lines = [f"Address {address.index}"]
        values = {
            "Type": address.address_type,
            "Line 1": address.line_1,
            "Line 2": address.line_2,
            "Town or city": address.town_or_city,
            "County": address.county,
            "Postcode": address.postcode,
            "Areas of law": address.areas_of_law,
            "Court types": address.court_types,
        }
        lines.extend(
            f"{label}: {_display_value(value)}" for label, value in values.items() if _has_value(value)
        )
        addresses.append("\n".join(lines))
    return "\n\n".join(addresses) if addresses else None


def _format_counter_service(submission: CourtSubmission) -> str | None:
    counter_service = submission.counter_service
    lines = [
        f"{label}: {_display_value(counter_service.get(key))}"
        for key, label in COUNTER_SERVICE_LABELS.items()
        if _has_value(counter_service.get(key))
    ]
    lines.extend(
        _format_time_line(label, counter_service.get(key))
        for key, label in (
            ("monday_to_friday", "Monday to Friday"),
            ("monday", "Monday"),
            ("tuesday", "Tuesday"),
            ("wednesday", "Wednesday"),
            ("thursday", "Thursday"),
            ("friday", "Friday"),
        )
        if _format_time_line(label, counter_service.get(key))
    )
    return "\n".join(lines) if lines else None


def _format_contacts(submission: CourtSubmission) -> str | None:
    contacts = []
    for contact in submission.contacts:
        lines = [f"Contact {contact.index}"]
        values = {
            "Description": contact.description,
            "Explanation": contact.explanation,
            "Phone": contact.phone,
            "Email": contact.email,
        }
        lines.extend(
            f"{label}: {_display_value(value)}" for label, value in values.items() if _has_value(value)
        )
        contacts.append("\n".join(lines))
    return "\n\n".join(contacts) if contacts else None


def _format_opening_hours(submission: CourtSubmission) -> str | None:
    opening_hours = []
    for opening_hours_set in submission.opening_hours:
        lines = [f"Opening hours {opening_hours_set.index}"]
        if _has_value(opening_hours_set.type):
            lines.append(f"Type: {_display_value(opening_hours_set.type)}")
        if opening_hours_set.same_monday_to_friday is not None:
            lines.append(
                "Same Monday to Friday: "
                f"{_display_value(opening_hours_set.same_monday_to_friday)}"
            )
        lines.extend(
            _format_time_line(label, value)
            for label, value in (
                ("Monday to Friday", opening_hours_set.monday_to_friday),
                ("Monday", opening_hours_set.monday),
                ("Tuesday", opening_hours_set.tuesday),
                ("Wednesday", opening_hours_set.wednesday),
                ("Thursday", opening_hours_set.thursday),
                ("Friday", opening_hours_set.friday),
            )
            if _format_time_line(label, value)
        )
        opening_hours.append("\n".join(lines))
    return "\n\n".join(opening_hours) if opening_hours else None


def _format_time_line(label: str, value: Any) -> str | None:
    if not value:
        return None
    if isinstance(value, dict):
        opening_time = value.get("open")
        closing_time = value.get("close")
        status = value.get("status")
    else:
        opening_time = value.open
        closing_time = value.close
        status = value.status

    if opening_time and closing_time:
        return f"{label}: {opening_time} to {closing_time}"
    if status:
        return f"{label}: {status.replace('_', ' ')}"
    return None


def _display_value(value: Any) -> str:
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, list):
        return "; ".join(_display_value(item) for item in value)
    return str(value)


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (list, tuple, set, dict)):
        return bool(value)
    return bool(str(value).strip())


def _highlight_decision_columns(worksheet: Worksheet, start_row: int) -> None:
    header_indexes = {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value in {"nsu_decision", "selected_source_row", "decision_notes", "decision_maker", "decision_date"}
    }
    for row_number in range(start_row, worksheet.max_row + 1):
        for column_index in header_indexes.values():
            worksheet.cell(row_number, column_index).fill = DECISION_FILL


def _write_table(
    workbook: Workbook,
    title: str,
    rows: list[list[Any]],
    *,
    wrap_text: bool = False,
    data_row_height: float | None = None,
) -> Worksheet:
    worksheet = workbook.create_sheet(title)
    for row in rows:
        worksheet.append(row)
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        width = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 12), 60)
    if wrap_text:
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if data_row_height is not None:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[row_number].height = data_row_height
    return worksheet


def _set_column_widths(worksheet: Worksheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
