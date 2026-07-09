"""Submitter read-only approval role outputs."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from fact_form_importer.cleaners.strings import null_if_empty_like
from fact_form_importer.models.court_submission import CourtSubmission

READ_ONLY_APPROVAL_ROLE = "read_only_approval"
JSON_NAME = "read_only_approval_users.json"
WORKBOOK_NAME = "read_only_approval_users.xlsx"
DEFAULT_EXCLUSIONS_PATH = Path("config/team_exclusions.json")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)


@dataclass(frozen=True)
class SubmitterOutputResult:
    json_path: Path
    workbook_path: Path
    user_count: int
    excluded_user_count: int


def write_submitter_outputs(
    submissions: list[CourtSubmission],
    output_path: Path,
    exclusions_path: Path = DEFAULT_EXCLUSIONS_PATH,
) -> SubmitterOutputResult:
    """Write read-only approval user JSON and workbook outputs."""

    output_path.mkdir(parents=True, exist_ok=True)
    payload = build_read_only_approval_users(submissions, exclusions_path)

    json_path = output_path / JSON_NAME
    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    workbook_path = output_path / WORKBOOK_NAME
    _write_workbook(payload, workbook_path)

    return SubmitterOutputResult(
        json_path=json_path,
        workbook_path=workbook_path,
        user_count=len(payload["users"]),
        excluded_user_count=len(payload["excluded_users"]),
    )


def build_read_only_approval_users(
    submissions: list[CourtSubmission],
    exclusions_path: Path = DEFAULT_EXCLUSIONS_PATH,
) -> dict[str, Any]:
    exclusions = load_team_exclusions(exclusions_path)
    by_email: dict[str, dict[str, Any]] = {}

    for submission in submissions:
        email = normalise_submitter_email(submission.source.submitter_email)
        if email is None:
            continue

        user = by_email.setdefault(
            email,
            {
                "email": email,
                "name": None,
                "source_row_numbers": [],
            },
        )
        if user["name"] is None:
            user["name"] = null_if_empty_like(submission.source.submitter_name)
        user["source_row_numbers"].append(submission.source.source_row_number)

    excluded_users = []
    users = []
    for email, user in sorted(by_email.items()):
        user["source_row_numbers"] = sorted(set(user["source_row_numbers"]))
        if email in exclusions:
            excluded_users.append(
                {
                    "email": email,
                    "reason": "configured_exclusion",
                    "name": user["name"],
                    "source_row_numbers": user["source_row_numbers"],
                }
            )
            continue
        users.append(user)

    return {
        "role": READ_ONLY_APPROVAL_ROLE,
        "users": users,
        "excluded_users": excluded_users,
    }


def load_team_exclusions(path: Path = DEFAULT_EXCLUSIONS_PATH) -> set[str]:
    if not path.exists():
        return set()

    payload = json.loads(path.read_text(encoding="utf-8"))
    configured = payload.get("exclude_from_read_only_approval_role", [])
    if not isinstance(configured, list):
        raise ValueError("team_exclusions.json field 'exclude_from_read_only_approval_role' must be a list")

    return {
        email
        for value in configured
        if (email := normalise_submitter_email(value)) is not None
    }


def normalise_submitter_email(value: object) -> str | None:
    cleaned = null_if_empty_like(value)
    if cleaned is None:
        return None
    return cleaned.strip().lower()


def _write_workbook(payload: dict[str, Any], path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Read only approval users"
    worksheet.append(["email", "name", "source_row_numbers"])
    for user in payload["users"]:
        worksheet.append(
            [
                user["email"],
                user["name"],
                ", ".join(str(row_number) for row_number in user["source_row_numbers"]),
            ]
        )

    excluded = workbook.create_sheet("Excluded users")
    excluded.append(["email", "name", "reason", "source_row_numbers"])
    for user in payload["excluded_users"]:
        excluded.append(
            [
                user["email"],
                user.get("name"),
                user["reason"],
                ", ".join(str(row_number) for row_number in user.get("source_row_numbers", [])),
            ]
        )

    for sheet in workbook.worksheets:
        _format_sheet(sheet)

    workbook.save(path)


def _format_sheet(worksheet: Any) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
