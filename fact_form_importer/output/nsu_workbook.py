"""NSU review workbook generation."""

from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from fact_form_importer.models.court_submission import CourtSubmission, OpeningTime
from fact_form_importer.output.duplicate_review import (
    duplicate_timestamp,
    format_source_date,
    group_duplicate_submissions,
    most_recent_duplicate_submission,
)
from fact_form_importer.validators.base import (
    COURT_SLUG_AUTO_REPAIRED,
    COURT_SLUG_SUGGESTED,
)
from fact_form_importer.validators.os_addresses import AddressVerification

WORKBOOK_NAME = "nsu_cleaned_review.xlsx"
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
ISSUE_EXPLANATIONS = {
    "COURT_SLUG_NORMALISED": (
        "The submitted court identifier was changed into a clean slug. This is usually non-blocking."
    ),
    "COURT_SLUG_AUTO_REPAIRED": (
        "The submitted court slug did not exist, but a very high-confidence FaCT search match was verified and used."
    ),
    "COURT_SLUG_SUGGESTED": (
        "FaCT search found a possible court slug match, but confidence was not high enough to auto-repair."
    ),
    "DUPLICATE_COURT_SLUG": (
        "More than one submitted row resolves to the same court slug. The importer does not choose or merge rows automatically."
    ),
    "INVALID_EMAIL": "An email value could not be parsed as a valid email address.",
    "INVALID_PHONE": "A phone value could not be parsed as a possible UK phone number.",
    "INVALID_POSTCODE": "A populated address postcode does not match the expected UK postcode format.",
    "INVALID_TIME": "An opening-hours value could not be parsed as a valid HH:MM time.",
    "MISSING_COURT_IDENTIFIER": "The row has business data but no usable court identifier.",
    "OPENING_HOURS_AMBIGUOUS": "Opening hours need review because the time values are invalid or ambiguous.",
    "POSTCODE_TYPO_REPAIRED": "An obvious O/0 typo in a postcode digit position was repaired.",
    "VOCAB_NO_MATCH": "A value does not match the configured controlled list.",
    "ADDRESS_OS_NORMALISED": "A unique high-confidence Ordnance Survey candidate was used to normalise an address.",
    "ADDRESS_OS_VERIFIED": "The submitted address matched a unique Ordnance Survey candidate without changes.",
    "ADDRESS_OS_REVIEW_REQUIRED": "Ordnance Survey returned address candidates, but none was safe to choose automatically.",
    "ADDRESS_OS_LOOKUP_UNAVAILABLE": "The FaCT/Ordnance Survey lookup was temporarily unavailable; no address change was made.",
}
ISSUE_ACTIONS = {
    "COURT_SLUG_NORMALISED": "Check only if the cleaned slug looks wrong.",
    "COURT_SLUG_AUTO_REPAIRED": "Check only if the repaired slug looks wrong.",
    "COURT_SLUG_SUGGESTED": "Review the suggested slug and correct the court slug if it is the right match.",
    "DUPLICATE_COURT_SLUG": "Review all rows in the duplicate group and decide whether to merge, discard, or correct them.",
    "INVALID_EMAIL": "Correct the email address or confirm it should be omitted.",
    "INVALID_PHONE": "Correct the phone number or confirm it should be omitted.",
    "INVALID_POSTCODE": "Correct the postcode before import.",
    "INVALID_TIME": "Correct the opening-hours time before import.",
    "MISSING_COURT_IDENTIFIER": "Add a valid court slug before import.",
    "OPENING_HOURS_AMBIGUOUS": "Review the opening-hours fields and correct the time values.",
    "POSTCODE_TYPO_REPAIRED": "Check only if the cleaned postcode looks wrong.",
    "VOCAB_NO_MATCH": "Map the value to an allowed option or confirm it needs a new controlled-list value.",
    "ADDRESS_OS_NORMALISED": "Check the OS-derived address if it looks unexpected.",
    "ADDRESS_OS_VERIFIED": "No action needed unless the submitted address appears wrong.",
    "ADDRESS_OS_REVIEW_REQUIRED": "Review the OS candidates and choose or correct the address before sending that address action.",
    "ADDRESS_OS_LOOKUP_UNAVAILABLE": "Run address verification again after FaCT/Ordnance Survey is available.",
}


def write_nsu_review_workbook(
    submissions: list[CourtSubmission],
    output_path: Path,
    summary: dict[str, Any],
    address_verifications: list[AddressVerification] | None = None,
) -> Path:
    output_path.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _add_summary_sheet(workbook, submissions, summary)
    _add_records_sheet(
        workbook,
        "Processed records",
        [submission for submission in submissions if submission.status in {"processed", "processed_with_warnings"}],
    )
    _add_records_sheet(
        workbook,
        "Needs human review",
        [submission for submission in submissions if submission.status == "needs_human_review"],
    )
    _add_records_sheet(
        workbook,
        "Failed records",
        [submission for submission in submissions if submission.status == "failed"],
    )
    _add_duplicate_courts_sheet(workbook, submissions)
    _add_court_slug_suggestions_sheet(workbook, submissions)
    _add_addresses_sheet(workbook, submissions)
    _add_address_verification_sheet(workbook, address_verifications or [])
    _add_contacts_sheet(workbook, submissions)
    _add_opening_hours_sheet(workbook, submissions)
    _add_issues_sheet(workbook, submissions)
    _add_submitter_users_sheet(workbook, submissions)

    path = output_path / WORKBOOK_NAME
    workbook.save(path)
    return path


def _add_summary_sheet(
    workbook: Workbook,
    submissions: list[CourtSubmission],
    summary: dict[str, Any],
) -> None:
    status_counts = Counter(submission.status for submission in submissions)
    issue_counts = Counter(
        issue.code
        for submission in submissions
        for issue in submission.issues
    )
    rows = [
        ["Metric", "Value"],
        ["Run ID", summary.get("run_id")],
        ["Source file", summary.get("source_file")],
        ["Workbook rows", summary.get("row_count")],
        ["Submissions", len(submissions)],
        ["Processed", status_counts["processed"]],
        ["Processed with warnings", status_counts["processed_with_warnings"]],
        ["Needs human review", status_counts["needs_human_review"]],
        ["Failed", status_counts["failed"]],
        ["Skipped empty rows", summary.get("skipped_count")],
        ["Duplicate slug groups", summary.get("duplicate_slug_group_count")],
        [
            "Duplicate affected records (included in needs human review)",
            summary.get("duplicate_slug_affected_record_count"),
        ],
        ["Address verification enabled", summary.get("address_verification_enabled")],
        ["Addresses checked against OS", summary.get("address_verification_count")],
        ["Unique postcode lookups", summary.get("address_verification_unique_postcode_lookups")],
        ["Addresses auto-normalised from OS", summary.get("address_verification_auto_normalised_count")],
        ["Addresses with ambiguous OS candidates", summary.get("address_verification_review_required_count")],
        ["Address actions held for review", summary.get("address_verification_action_blocking_count")],
        [],
        ["Issue code", "Count"],
    ]
    rows.extend([code, count] for code, count in sorted(issue_counts.items()))
    _write_sheet(workbook, "Summary", rows, freeze=False, autofilter=False)


def _add_records_sheet(
    workbook: Workbook,
    title: str,
    submissions: list[CourtSubmission],
) -> None:
    rows = [
        [
            "source_row_number",
            "court_slug",
            "status",
            "issue_count",
            "issue_codes",
            "review_reason",
            "suggested_next_action",
            "court_slug_raw",
            "translation_phone",
            "translation_email",
            "address_count",
            "contact_count",
            "opening_hours_count",
            "suggested_slug",
            "suggested_court_name",
            "suggestion_confidence",
        ]
    ]
    for submission in submissions:
        suggestion = _court_slug_suggestion(submission)
        rows.append(
            [
                submission.source.source_row_number,
                submission.court_slug,
                submission.status,
                len(submission.issues),
                ", ".join(sorted({issue.code for issue in submission.issues})),
                _review_reason(submission),
                _suggested_next_action(submission),
                submission.court_slug_raw,
                submission.translation_phone,
                submission.translation_email,
                len(submission.addresses),
                len(submission.contacts),
                len(submission.opening_hours),
                suggestion.get("suggested_slug"),
                suggestion.get("suggested_court_name"),
                suggestion.get("confidence"),
            ]
        )

    _write_sheet(workbook, title, rows)


def _add_duplicate_courts_sheet(workbook: Workbook, submissions: list[CourtSubmission]) -> None:
    by_slug = group_duplicate_submissions(submissions)

    rows = [
        [
            "court_slug",
            "record_count",
            "source_row_numbers",
            "candidate_most_recent_row",
            "candidate_most_recent_date",
            "submitter_names",
            "submitter_emails",
            "completion_times",
            "start_times",
            "last_modified_times",
            "statuses",
            "raw_slug_values",
            "source_rows_with_dates",
        ]
    ]
    for slug, matching_submissions in sorted(by_slug.items()):
        most_recent = most_recent_duplicate_submission(matching_submissions)
        candidate_timestamp = duplicate_timestamp(most_recent) if most_recent else None
        rows.append(
            [
                slug,
                len(matching_submissions),
                ", ".join(str(submission.source.source_row_number) for submission in matching_submissions),
                most_recent.source.source_row_number if most_recent else None,
                candidate_timestamp.display_value if candidate_timestamp else None,
                " | ".join(
                    sorted(
                        {
                            submission.source.submitter_name
                            for submission in matching_submissions
                            if submission.source.submitter_name
                        }
                    )
                ),
                " | ".join(
                    sorted(
                        {
                            submission.source.submitter_email
                            for submission in matching_submissions
                            if submission.source.submitter_email
                        }
                    )
                ),
                " | ".join(
                    _dedupe_preserving_order(
                        format_source_date(submission.source.completion_time)
                        for submission in matching_submissions
                    )
                ),
                " | ".join(
                    _dedupe_preserving_order(
                        format_source_date(submission.source.start_time)
                        for submission in matching_submissions
                    )
                ),
                " | ".join(
                    _dedupe_preserving_order(
                        format_source_date(submission.source.last_modified_time)
                        for submission in matching_submissions
                    )
                ),
                ", ".join(sorted({submission.status for submission in matching_submissions})),
                " | ".join(sorted({str(submission.court_slug_raw) for submission in matching_submissions})),
                " | ".join(
                    _duplicate_submission_summary(submission)
                    for submission in sorted(
                        matching_submissions,
                        key=lambda item: item.source.source_row_number,
                    )
                ),
            ]
        )

    _write_sheet(workbook, "Duplicate courts", rows)


def _add_court_slug_suggestions_sheet(workbook: Workbook, submissions: list[CourtSubmission]) -> None:
    rows = [
        [
            "source_row_number",
            "status",
            "issue_code",
            "raw_slug_value",
            "current_court_slug",
            "submitted_slug",
            "suggested_slug",
            "suggested_court_name",
            "confidence",
            "query",
            "reason",
        ]
    ]
    for submission in submissions:
        for issue in submission.issues:
            if issue.code not in {COURT_SLUG_AUTO_REPAIRED, COURT_SLUG_SUGGESTED}:
                continue

            suggestion = issue.cleaned_value if isinstance(issue.cleaned_value, dict) else {}
            rows.append(
                [
                    submission.source.source_row_number,
                    submission.status,
                    issue.code,
                    submission.court_slug_raw,
                    submission.court_slug,
                    suggestion.get("submitted_slug"),
                    suggestion.get("suggested_slug"),
                    suggestion.get("suggested_court_name"),
                    suggestion.get("confidence"),
                    suggestion.get("query"),
                    suggestion.get("reason"),
                ]
            )

    _write_sheet(workbook, "Court slug suggestions", rows)


def _duplicate_submission_summary(submission: CourtSubmission) -> str:
    timestamp = duplicate_timestamp(submission)
    date = timestamp.display_value if timestamp else "no date"
    submitter = submission.source.submitter_name or submission.source.submitter_email or "unknown submitter"
    return f"row {submission.source.source_row_number}: {date}, {submitter}"


def _dedupe_preserving_order(values: Iterable[str | None]) -> list[str]:
    seen = set()
    deduped = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        deduped.append(value)
    return deduped


def _add_addresses_sheet(workbook: Workbook, submissions: list[CourtSubmission]) -> None:
    rows = [
        [
            "source_row_number",
            "court_slug",
            "status",
            "address_index",
            "address_type",
            "line_1",
            "line_2",
            "town_or_city",
            "county",
            "postcode",
            "areas_of_law",
            "court_types",
        ]
    ]
    for submission in submissions:
        for address in submission.addresses:
            rows.append(
                [
                    submission.source.source_row_number,
                    submission.court_slug,
                    submission.status,
                    address.index,
                    address.address_type,
                    address.line_1,
                    address.line_2,
                    address.town_or_city,
                    address.county,
                    address.postcode,
                    "; ".join(address.areas_of_law),
                    "; ".join(address.court_types),
                ]
            )

    _write_sheet(workbook, "Cleaned addresses", rows)


def _add_address_verification_sheet(
    workbook: Workbook,
    verifications: list[AddressVerification],
) -> None:
    rows = [
        [
            "source_row_number",
            "court_slug",
            "address_index",
            "verification_status",
            "message",
            "match_score",
            "score_margin",
            "match_type",
            "selected_uprn",
            "original_address",
            "proposed_address",
            "llm_suggested_uprn",
            "llm_confidence",
            "llm_needs_human_review",
            "llm_reason",
            "os_candidates",
        ]
    ]
    for verification in verifications:
        suggestion = verification.llm_suggestion or {}
        rows.append(
            [
                verification.source_row_number,
                verification.court_slug,
                verification.address_index,
                verification.status,
                verification.message,
                verification.match_score,
                verification.score_margin,
                verification.match_type,
                verification.selected_candidate.uprn if verification.selected_candidate else None,
                _cell_value(verification.original_address),
                _cell_value(verification.proposed_address),
                suggestion.get("uprn"),
                suggestion.get("confidence"),
                suggestion.get("needs_human_review"),
                suggestion.get("reason"),
                _cell_value([candidate.as_dict() for candidate in verification.candidates]),
            ]
        )
    _write_sheet(workbook, "Address verification", rows)


def _add_contacts_sheet(workbook: Workbook, submissions: list[CourtSubmission]) -> None:
    rows = [
        [
            "source_row_number",
            "court_slug",
            "status",
            "contact_index",
            "description",
            "explanation",
            "phone",
            "email",
        ]
    ]
    for submission in submissions:
        for contact in submission.contacts:
            rows.append(
                [
                    submission.source.source_row_number,
                    submission.court_slug,
                    submission.status,
                    contact.index,
                    contact.description,
                    contact.explanation,
                    contact.phone,
                    contact.email,
                ]
            )

    _write_sheet(workbook, "Cleaned contacts", rows)


def _add_opening_hours_sheet(workbook: Workbook, submissions: list[CourtSubmission]) -> None:
    rows = [
        [
            "source_row_number",
            "court_slug",
            "status",
            "opening_hours_index",
            "type",
            "same_monday_to_friday",
            "day",
            "open",
            "close",
            "time_status",
            "time_issue_codes",
        ]
    ]
    for submission in submissions:
        for opening_hours in submission.opening_hours:
            for day, time_value in _opening_time_values(opening_hours):
                rows.append(
                    [
                        submission.source.source_row_number,
                        submission.court_slug,
                        submission.status,
                        opening_hours.index,
                        opening_hours.type,
                        opening_hours.same_monday_to_friday,
                        day,
                        time_value.open if time_value else None,
                        time_value.close if time_value else None,
                        time_value.status if time_value else None,
                        ", ".join(issue.code for issue in time_value.issues) if time_value else None,
                    ]
                )

    _write_sheet(workbook, "Cleaned opening hours", rows)


def _add_issues_sheet(workbook: Workbook, submissions: list[CourtSubmission]) -> None:
    rows = [
        [
            "source_row_number",
            "court_slug",
            "field",
            "code",
            "severity",
            "message",
            "plain_english_meaning",
            "suggested_next_action",
            "raw_value",
            "cleaned_value",
            "suggested_slug",
            "suggested_court_name",
            "suggestion_confidence",
        ]
    ]
    for submission in submissions:
        for issue in submission.issues:
            rows.append(
                [
                    submission.source.source_row_number,
                    submission.court_slug,
                    issue.field,
                    issue.code,
                    issue.severity,
                    issue.message,
                    ISSUE_EXPLANATIONS.get(issue.code, "No explanation configured for this issue code."),
                    ISSUE_ACTIONS.get(issue.code, "Review the field value."),
                    _cell_value(issue.raw_value),
                    _cell_value(issue.cleaned_value),
                    _issue_suggestion_value(issue, "suggested_slug"),
                    _issue_suggestion_value(issue, "suggested_court_name"),
                    _issue_suggestion_value(issue, "confidence"),
                ]
            )

    _write_sheet(workbook, "Issues", rows)


def _add_submitter_users_sheet(workbook: Workbook, submissions: list[CourtSubmission]) -> None:
    by_submitter: dict[tuple[Any, Any], list[CourtSubmission]] = defaultdict(list)
    for submission in submissions:
        key = (submission.source.submitter_email, submission.source.submitter_name)
        by_submitter[key].append(submission)

    rows = [
        [
            "submitter_email",
            "submitter_name",
            "submission_count",
            "source_row_numbers",
            "statuses",
        ]
    ]
    for (email, name), matching_submissions in sorted(by_submitter.items(), key=lambda item: str(item[0])):
        rows.append(
            [
                email,
                name,
                len(matching_submissions),
                ", ".join(str(submission.source.source_row_number) for submission in matching_submissions),
                ", ".join(sorted({submission.status for submission in matching_submissions})),
            ]
        )

    _write_sheet(workbook, "Submitter users", rows)


def _opening_time_values(opening_hours: Any) -> Iterable[tuple[str, OpeningTime | None]]:
    yield "Monday to Friday", opening_hours.monday_to_friday
    yield "Monday", opening_hours.monday
    yield "Tuesday", opening_hours.tuesday
    yield "Wednesday", opening_hours.wednesday
    yield "Thursday", opening_hours.thursday
    yield "Friday", opening_hours.friday


def _review_reason(submission: CourtSubmission) -> str:
    if not submission.issues:
        return "No validation issues."

    issue_codes = sorted({issue.code for issue in submission.issues})
    if submission.status == "processed_with_warnings":
        return "Only non-blocking warnings were found: " + _issue_explanations(
            issue_codes,
            submission,
        )

    return _issue_explanations(issue_codes, submission)


def _suggested_next_action(submission: CourtSubmission) -> str:
    if not submission.issues:
        return "No action required."

    issue_codes = sorted({issue.code for issue in submission.issues})
    return " ".join(ISSUE_ACTIONS.get(code, "Review the field value.") for code in issue_codes)


def _issue_explanations(issue_codes: list[str], submission: CourtSubmission) -> str:
    return " ".join(_issue_explanation(code, submission) for code in issue_codes)


def _issue_explanation(code: str, submission: CourtSubmission) -> str:
    matching_issues = [issue for issue in submission.issues if issue.code == code]
    base = ISSUE_EXPLANATIONS.get(code, "No explanation configured for this issue code.")

    if code == "VOCAB_NO_MATCH":
        details = []
        for issue in matching_issues:
            details.append(
                f"{issue.field} value {_cell_value(issue.raw_value)!r} did not match"
            )
        return f"{code}: {base} " + "; ".join(details) + "."

    if code in {"INVALID_EMAIL", "INVALID_PHONE", "INVALID_POSTCODE", "INVALID_TIME"}:
        details = []
        for issue in matching_issues:
            details.append(f"{issue.field} value {_cell_value(issue.raw_value)!r}")
        return f"{code}: {base} Affected value(s): " + "; ".join(details) + "."

    if code in {COURT_SLUG_AUTO_REPAIRED, COURT_SLUG_SUGGESTED}:
        details = []
        for issue in matching_issues:
            suggestion = issue.cleaned_value if isinstance(issue.cleaned_value, dict) else {}
            details.append(
                f"{suggestion.get('suggested_slug')} "
                f"({suggestion.get('suggested_court_name')}, confidence {suggestion.get('confidence')})"
            )
        return f"{code}: {base} Suggested match(es): " + "; ".join(details) + "."

    return f"{code}: {base}"


def _court_slug_suggestion(submission: CourtSubmission) -> dict[str, Any]:
    for issue in submission.issues:
        if issue.code in {COURT_SLUG_AUTO_REPAIRED, COURT_SLUG_SUGGESTED} and isinstance(issue.cleaned_value, dict):
            return issue.cleaned_value
    return {}


def _issue_suggestion_value(issue: Any, key: str) -> Any:
    if isinstance(issue.cleaned_value, dict):
        return issue.cleaned_value.get(key)
    return None


def _write_sheet(
    workbook: Workbook,
    title: str,
    rows: list[list[Any]],
    freeze: bool = True,
    autofilter: bool = True,
) -> Worksheet:
    worksheet = workbook.create_sheet(title=title)
    for row in rows:
        worksheet.append([_cell_value(value) for value in row])

    if rows:
        for cell in worksheet[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

    if freeze:
        worksheet.freeze_panes = "A2"

    if autofilter and rows and rows[0]:
        worksheet.auto_filter.ref = worksheet.dimensions

    _set_column_widths(worksheet)
    return worksheet


def _set_column_widths(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _cell_value(value: Any) -> Any:
    if isinstance(value, BaseModel):
        return json.dumps(value.model_dump(mode="json"), ensure_ascii=False, sort_keys=True)

    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)

    return value
