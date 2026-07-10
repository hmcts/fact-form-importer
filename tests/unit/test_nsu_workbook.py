from openpyxl import load_workbook

from fact_form_importer.models.court_submission import (
    Address,
    ContactDetail,
    CourtSubmission,
    OpeningHoursSet,
    OpeningTime,
)
from fact_form_importer.models.issues import Issue
from fact_form_importer.models.source import SourceMetadata
from fact_form_importer.output.nsu_workbook import WORKBOOK_NAME, write_nsu_review_workbook


def test_write_nsu_review_workbook_creates_expected_review_tabs(tmp_path):
    submissions = [
        _submission(
            court_slug="processed-court",
            status="processed",
            row_number=2,
            addresses=[Address(index=1, address_type="Visit", postcode="SW1A 1AA")],
            contacts=[ContactDetail(index=1, description="Enquiries", phone="020 7946 0000")],
            opening_hours=[
                OpeningHoursSet(
                    index=1,
                    type="Court open",
                    monday_to_friday=OpeningTime(
                        open="09:00",
                        close="17:00",
                        status="valid_time",
                    ),
                )
            ],
        ),
        _submission(
            court_slug="duplicate-court",
            status="needs_human_review",
            row_number=3,
            issue_code="DUPLICATE_COURT_SLUG",
            completion_time="2026-07-01 09:00",
            submitter_name="Earlier Submitter",
            submitter_email="earlier@example.com",
        ),
        _submission(
            court_slug="duplicate-court",
            status="needs_human_review",
            row_number=4,
            issue_code="DUPLICATE_COURT_SLUG",
            completion_time="2026-07-02 10:00",
            submitter_name="Later Submitter",
            submitter_email="later@example.com",
        ),
        _submission(
            court_slug=None,
            status="failed",
            row_number=5,
            issue_code="MISSING_COURT_IDENTIFIER",
        ),
    ]
    summary = {
        "run_id": "run-1",
        "source_file": "source.xlsx",
        "row_count": 5,
        "skipped_count": 0,
        "duplicate_slug_group_count": 1,
        "duplicate_slug_affected_record_count": 2,
    }

    path = write_nsu_review_workbook(submissions, tmp_path, summary)

    assert path == tmp_path / WORKBOOK_NAME
    workbook = load_workbook(path)
    assert workbook.sheetnames == [
        "Summary",
        "Processed records",
        "Needs human review",
        "Failed records",
        "Duplicate courts",
        "Court slug suggestions",
        "Cleaned addresses",
        "Cleaned contacts",
        "Cleaned opening hours",
        "Issues",
        "Submitter users",
    ]
    assert workbook["Processed records"].freeze_panes == "A2"
    assert workbook["Processed records"].auto_filter.ref is not None
    assert workbook["Processed records"]["B2"].value == "processed-court"
    assert workbook["Needs human review"]["F2"].value.startswith("DUPLICATE_COURT_SLUG:")
    assert "Review all rows" in workbook["Needs human review"]["G2"].value
    assert (
        workbook["Summary"]["A12"].value
        == "Duplicate affected records (included in needs human review)"
    )
    assert workbook["Duplicate courts"]["A2"].value == "duplicate-court"
    assert workbook["Duplicate courts"]["D2"].value == 4
    assert workbook["Duplicate courts"]["E2"].value == "2026-07-02 10:00"
    assert workbook["Duplicate courts"]["F2"].value == "Earlier Submitter | Later Submitter"
    assert workbook["Duplicate courts"]["G2"].value == "earlier@example.com | later@example.com"
    assert "row 3: 2026-07-01 09:00, Earlier Submitter" in workbook["Duplicate courts"]["M2"].value
    assert "row 4: 2026-07-02 10:00, Later Submitter" in workbook["Duplicate courts"]["M2"].value
    assert workbook["Cleaned addresses"]["J2"].value == "SW1A 1AA"
    assert workbook["Cleaned contacts"]["E2"].value == "Enquiries"
    assert workbook["Cleaned opening hours"]["H2"].value == "09:00"
    assert workbook["Issues"]["D2"].value == "DUPLICATE_COURT_SLUG"
    assert "More than one submitted row" in workbook["Issues"]["G2"].value
    assert "Review all rows" in workbook["Issues"]["H2"].value


def test_nsu_workbook_review_reason_names_vocab_field_and_raw_value(tmp_path):
    submission = _submission(
        court_slug="review-court",
        status="needs_human_review",
        row_number=2,
        issue=Issue(
            field="contacts[1].description",
            code="VOCAB_NO_MATCH",
            severity="warning",
            message="Value does not match vocabulary 'contact_description_types'",
            raw_value="General help",
        ),
    )

    path = write_nsu_review_workbook([submission], tmp_path, {"run_id": "run-1"})

    workbook = load_workbook(path)
    review_reason = workbook["Needs human review"]["F2"].value
    assert "contacts[1].description" in review_reason
    assert "General help" in review_reason


def test_nsu_workbook_writes_court_slug_suggestions_tab(tmp_path):
    submission = _submission(
        court_slug="shrewsbury-crown-court",
        status="processed_with_warnings",
        row_number=2,
        issue=Issue(
            field="court_slug",
            code="COURT_SLUG_AUTO_REPAIRED",
            severity="warning",
            message="Court slug was auto-repaired",
            raw_value="shrewsbury.crowncourt",
            cleaned_value={
                "submitted_slug": "shrewsburycrowncourt",
                "suggested_slug": "shrewsbury-crown-court",
                "suggested_court_name": "Shrewsbury Crown Court",
                "confidence": 1.0,
                "query": "shrewsbury crown court",
                "reason": "Best match from FaCT court name search",
            },
        ),
    )

    path = write_nsu_review_workbook([submission], tmp_path, {"run_id": "run-1"})

    workbook = load_workbook(path)
    assert workbook["Processed records"]["N2"].value == "shrewsbury-crown-court"
    assert workbook["Processed records"]["O2"].value == "Shrewsbury Crown Court"
    assert workbook["Processed records"]["P2"].value == 1
    assert workbook["Court slug suggestions"]["G2"].value == "shrewsbury-crown-court"
    assert workbook["Court slug suggestions"]["H2"].value == "Shrewsbury Crown Court"


def _submission(
    court_slug,
    status,
    row_number,
    addresses=None,
    contacts=None,
    opening_hours=None,
    issue_code=None,
    issue=None,
    completion_time=None,
    start_time=None,
    last_modified_time=None,
    submitter_email="submitter@example.com",
    submitter_name="Submitter",
):
    issues = []
    if issue:
        issues.append(issue)
    if issue_code:
        issues.append(
            Issue(
                field="court_slug",
                code=issue_code,
                severity="warning",
                message="Test issue",
            )
        )

    return CourtSubmission(
        source=SourceMetadata(
            source_row_number=row_number,
            completion_time=completion_time,
            start_time=start_time,
            last_modified_time=last_modified_time,
            submitter_email=submitter_email,
            submitter_name=submitter_name,
        ),
        court_slug_raw=court_slug,
        court_slug=court_slug,
        status=status,
        addresses=addresses or [],
        contacts=contacts or [],
        opening_hours=opening_hours or [],
        issues=issues,
    )
