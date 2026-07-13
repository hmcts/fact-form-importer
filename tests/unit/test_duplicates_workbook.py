from openpyxl import load_workbook

from fact_form_importer.models.court_submission import (
    Address,
    ContactDetail,
    CourtSubmission,
    OpeningHoursSet,
    OpeningTime,
)
from fact_form_importer.models.issues import Issue
from fact_form_importer.models.source import SourceMetadata
from fact_form_importer.output.duplicates_workbook import (
    WORKBOOK_NAME,
    write_duplicate_review_workbook,
)


def test_duplicate_review_workbook_groups_forms_and_marks_only_a_date_based_candidate(tmp_path):
    submissions = [
        _duplicate("example-court", 2, completion_time="2026-07-01 09:00", submitter_name="Earlier"),
        _duplicate(
            "example-court",
            3,
            completion_time="2026-07-02 10:00",
            submitter_name="Later",
            raw={"H": "Yes", "AA": "Visit", "AB": "1 Example Street"},
            facilities={"accessible_parking": True},
            addresses=[
                Address(
                    index=1,
                    address_type="Visit",
                    line_1="1 Example Street",
                    postcode="SW1A 1AA",
                )
            ],
            contacts=[
                ContactDetail(index=1, description="Enquiries", phone="020 7946 0000")
            ],
            opening_hours=[
                OpeningHoursSet(
                    index=1,
                    type="Court open",
                    same_monday_to_friday=True,
                    monday_to_friday=OpeningTime(
                        open="09:00",
                        close="17:00",
                        status="valid_time",
                    ),
                )
            ],
        ),
        _duplicate("no-date-court", 4, submitter_name="No Date One"),
        _duplicate("no-date-court", 5, submitter_name="No Date Two"),
    ]

    path = write_duplicate_review_workbook(
        submissions,
        tmp_path,
        {"run_id": "run-1", "source_file": "forms.xlsx"},
    )

    assert path == tmp_path / WORKBOOK_NAME
    workbook = load_workbook(path)
    assert workbook.sheetnames == [
        "Summary",
        "Duplicate group overview",
        "Duplicate form rows",
        "Duplicate form data",
        "Decision log",
    ]

    overview = workbook["Duplicate group overview"]
    assert overview.freeze_panes == "A2"
    assert overview.auto_filter.ref is not None
    assert overview["A2"].value == "D001"
    assert overview["B2"].value == "example-court"
    assert overview["E2"].value == 3
    assert overview["F2"].value == "2026-07-02 10:00"
    assert overview["G2"].value == "Completion time"
    assert overview["E3"].value is None
    assert overview["G3"].value == "No timestamp available"

    form_rows = workbook["Duplicate form rows"]
    rows = {form_rows.cell(row, 4).value: row for row in range(2, form_rows.max_row + 1)}
    assert form_rows.cell(rows[3], 3).value == "Yes - date-based candidate"
    assert form_rows.cell(rows[2], 3).value == "No"
    assert form_rows.cell(rows[3], 1).value == "D001"
    assert form_rows.cell(rows[3], 11).value == "Later"

    form_data = workbook["Duplicate form data"]
    headers = {cell.value: cell.column for cell in form_data[1]}
    data_rows = {
        form_data.cell(row, headers["source_row_number"]).value: row
        for row in range(2, form_data.max_row + 1)
    }
    assert form_data.freeze_panes == "A2"
    assert form_data.auto_filter.ref is not None
    assert form_data.cell(data_rows[3], headers["submitted_form_answers"]).alignment.wrap_text is True
    assert form_data.cell(data_rows[3], headers["date_based_candidate"]).value == "Yes - date-based candidate"
    assert "Is there accessible parking?: Yes" in form_data.cell(
        data_rows[3], headers["submitted_form_answers"]
    ).value
    assert "Accessible parking: Yes" in form_data.cell(
        data_rows[3], headers["cleaned_facilities"]
    ).value
    assert "Address 1" in form_data.cell(data_rows[3], headers["cleaned_addresses"]).value
    assert "Contact 1" in form_data.cell(data_rows[3], headers["cleaned_contact_details"]).value
    assert "Monday to Friday: 09:00 to 17:00" in form_data.cell(
        data_rows[3], headers["cleaned_court_opening_hours"]
    ).value

    decision_log = workbook["Decision log"]
    assert decision_log["A2"].value == "D001"
    assert decision_log["C2"].value == 3
    assert decision_log["E2"].value is None


def test_duplicate_review_workbook_uses_last_modified_then_start_time_when_completion_is_blank(tmp_path):
    submissions = [
        _duplicate("example-court", 2, last_modified_time="2026-07-03 10:00"),
        _duplicate("example-court", 3, start_time="2026-07-02 10:00"),
    ]

    path = write_duplicate_review_workbook(submissions, tmp_path, {"run_id": "run-1"})
    workbook = load_workbook(path)
    overview = workbook["Duplicate group overview"]

    assert overview["E2"].value == 2
    assert overview["F2"].value == "2026-07-03 10:00"
    assert overview["G2"].value == "Last modified time"


def _duplicate(court_slug, row_number, **source_values):
    raw = source_values.pop("raw", {})
    facilities = source_values.pop("facilities", {})
    addresses = source_values.pop("addresses", [])
    contacts = source_values.pop("contacts", [])
    opening_hours = source_values.pop("opening_hours", [])
    return CourtSubmission(
        source=SourceMetadata(source_row_number=row_number, **source_values),
        court_slug=court_slug,
        court_slug_raw=court_slug,
        status="needs_human_review",
        raw=raw,
        facilities=facilities,
        addresses=addresses,
        contacts=contacts,
        opening_hours=opening_hours,
        issues=[
            Issue(
                field="court_slug",
                code="DUPLICATE_COURT_SLUG",
                severity="warning",
                message="Duplicate court slug",
            )
        ],
    )
