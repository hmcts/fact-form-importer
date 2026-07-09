import json

from openpyxl import load_workbook

from fact_form_importer.models.court_submission import CourtSubmission
from fact_form_importer.models.source import SourceMetadata
from fact_form_importer.output.submitters import (
    WORKBOOK_NAME,
    build_read_only_approval_users,
    write_submitter_outputs,
)


def test_build_read_only_approval_users_deduplicates_and_excludes(tmp_path):
    exclusions_path = tmp_path / "team_exclusions.json"
    exclusions_path.write_text(
        json.dumps(
            {
                "exclude_from_read_only_approval_role": [
                    "TEAM.Member@Justice.Gov.UK",
                ]
            }
        ),
        encoding="utf-8",
    )
    submissions = [
        _submission(2, "Person@example.gov.uk", "Person Example"),
        _submission(50, " person@example.gov.uk ", "Ignored Later Name"),
        _submission(51, "team.member@justice.gov.uk", "Team Member"),
        _submission(52, None, "No Email"),
    ]

    payload = build_read_only_approval_users(submissions, exclusions_path)

    assert payload["role"] == "read_only_approval"
    assert payload["users"] == [
        {
            "email": "person@example.gov.uk",
            "name": "Person Example",
            "source_row_numbers": [2, 50],
        }
    ]
    assert payload["excluded_users"] == [
        {
            "email": "team.member@justice.gov.uk",
            "reason": "configured_exclusion",
            "name": "Team Member",
            "source_row_numbers": [51],
        }
    ]


def test_write_submitter_outputs_writes_json_and_workbook(tmp_path):
    exclusions_path = tmp_path / "team_exclusions.json"
    exclusions_path.write_text('{"exclude_from_read_only_approval_role": []}', encoding="utf-8")
    submissions = [_submission(2, "person@example.gov.uk", "Person Example")]

    result = write_submitter_outputs(submissions, tmp_path, exclusions_path)

    assert result.user_count == 1
    assert result.excluded_user_count == 0
    assert result.json_path.exists()
    assert result.workbook_path == tmp_path / WORKBOOK_NAME

    payload = json.loads(result.json_path.read_text())
    assert payload["users"][0]["email"] == "person@example.gov.uk"

    workbook = load_workbook(result.workbook_path)
    assert workbook.sheetnames == ["Read only approval users", "Excluded users"]
    assert workbook["Read only approval users"]["A2"].value == "person@example.gov.uk"
    assert workbook["Read only approval users"]["C2"].value == "2"


def _submission(row_number, email, name):
    return CourtSubmission(
        source=SourceMetadata(
            source_row_number=row_number,
            submitter_email=email,
            submitter_name=name,
        )
    )
